from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles.differential import DifferentialStyle

# ── Palette ────────────────────────────────────────────────────
C_DARK      = "1A1A2E"   # navy header bg
C_ACCENT    = "E94560"   # red accent
C_BLUE      = "0F3460"   # deep blue
C_MID       = "16213E"   # mid navy
C_HEADER_TXT= "FFFFFF"
C_ROW_EVEN  = "F0F4FF"
C_ROW_ODD   = "FFFFFF"
C_BORDER    = "C5CAE9"
C_YELLOW    = "FFF9C4"
C_GREEN_BG  = "E8F5E9"
C_RED_BG    = "FFEBEE"
C_ORANGE_BG = "FFF3E0"

def thin_border(color=C_BORDER):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def header_cell(ws, row, col, text, bg=C_DARK, fg=C_HEADER_TXT, width=None):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=True, color=fg, size=10, name="Arial")
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin_border("FFFFFF")
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c

def data_row(ws, row, values, even=True):
    bg = C_ROW_EVEN if even else C_ROW_ODD
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(size=9, name="Arial", color="333333")
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = thin_border()

def freeze_and_filter(ws, row=2):
    ws.freeze_panes = ws.cell(row=row, column=1)
    ws.auto_filter.ref = ws.dimensions

def title_block(ws, title, subtitle, color=C_DARK):
    ws.row_dimensions[1].height = 36
    ws.merge_cells("A1:Z1")
    c = ws["A1"]
    c.value = title
    c.font = Font(bold=True, size=14, color="FFFFFF", name="Arial")
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = thin_border(color)

    ws.row_dimensions[2].height = 20
    ws.merge_cells("A2:Z2")
    c2 = ws["A2"]
    c2.value = subtitle
    c2.font = Font(italic=True, size=9, color="888888", name="Arial")
    c2.fill = PatternFill("solid", fgColor="FAFAFA")
    c2.alignment = Alignment(horizontal="left", vertical="center")


wb = Workbook()
wb.remove(wb.active)


# ══════════════════════════════════════════════════════════════
# 1. HOJA: Contexto
# ══════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("Contexto")
ws1.sheet_properties.tabColor = "0F3460"
ws1.sheet_view.showGridLines = False

title_block(ws1, "📋  CONTEXTO DE CONVERSACIONES", 
            "Historial activo por conversation_id — leído/escrito por el bot en cada mensaje", C_BLUE)

COLS_CTX = [
    ("conversation_id",  28, "ID único por canal+usuario  ej: whatsapp:34600123456"),
    ("from",             18, "ID/teléfono del remitente en el canal"),
    ("channel",          14, "whatsapp | instagram | facebook | test"),
    ("user_name",        20, "Nombre visible del usuario"),
    ("stage",            18, "Etapa actual: inicio | triaje | recopilando_nombre | recopilando_telefono | mostrando_slots | confirmando_cita | agendado | escalar"),
    ("history",          40, "JSON array con los últimos 20 turnos de conversación"),
    ("metadata",         40, "JSON: nombre, teléfono, edad, condiciones, tratamientos_interes, aptitudVerificada, slotPropuesto"),
    ("message_count",    14, "Nº total de mensajes del paciente en esta conv."),
    ("last_intencion",   18, "CONSULTA | AGENDAR | ESCALAR_HUMANO | CONFIRMACION_CITA | OTRO"),
    ("last_servicio",    22, "Último tratamiento mencionado"),
    ("cita_confirmada",  16, "TRUE / FALSE"),
    ("calendar_event_id",22, "ID del evento creado en Google Calendar"),
    ("escalar_humano",   16, "TRUE / FALSE"),
    ("last_activity",    22, "ISO timestamp de última actividad"),
    ("patient_id",       26, "FK → Hoja Pacientes"),
]

ROW_H = 3
ws1.row_dimensions[ROW_H].height = 38
for i, (col_name, width, tooltip) in enumerate(COLS_CTX, 1):
    bg = C_DARK if i % 2 == 0 else C_BLUE
    header_cell(ws1, ROW_H, i, col_name, bg=bg, width=width)

sample_data_ctx = [
    ["whatsapp:34600000001","34600000001","whatsapp","María García","confirmando_cita",
     '[{"role":"user","content":"Quiero botox"},{"role":"assistant","content":"Perfecto! ¿Cuál es tu nombre?"}]',
     '{"nombre":"María García","telefono":"600000001","edad":32,"condiciones":[],"tratamientos_interes":["Botox facial"],"aptitudVerificada":true,"slotPropuesto":{"inicio":"2026-05-06T10:00:00Z","label":"martes 6 de mayo a las 10:00"}}',
     "4","AGENDAR","Botox facial","FALSE","","FALSE","2026-04-29T10:32:00Z","PAC-1714300000-AB1CD"],
    ["instagram:987654321","987654321","instagram","Carlos López","triaje",
     '[{"role":"user","content":"Hola quiero info de PRP"}]',
     '{"nombre":null,"telefono":null,"edad":null,"condiciones":[],"tratamientos_interes":["PRP"],"aptitudVerificada":false}',
     "1","AGENDAR","PRP Plasma Rico Plaquetas","FALSE","","FALSE","2026-04-29T11:05:00Z",""],
    ["facebook:111222333","111222333","facebook","Ana Martínez","agendado",
     '[]',
     '{"nombre":"Ana Martínez","telefono":"610000002","edad":28,"condiciones":[],"tratamientos_interes":["Limpieza Facial Profunda"],"aptitudVerificada":true}',
     "7","CONFIRMACION_CITA","Limpieza Facial Profunda","TRUE","cal_event_xyz789","FALSE","2026-04-28T16:00:00Z","PAC-1714200000-XY2ZW"],
]

for r_idx, row in enumerate(sample_data_ctx, ROW_H + 1):
    ws1.row_dimensions[r_idx].height = 40
    data_row(ws1, r_idx, row, even=(r_idx % 2 == 0))

# Conditional formatting for stage column (E)
stage_colors = {
    "agendado":             "C8E6C9",
    "confirmando_cita":     "B3E5FC",
    "mostrando_slots":      "E1BEE7",
    "recopilando_telefono": "FFF9C4",
    "recopilando_nombre":   "FFF9C4",
    "triaje":               "FFE0B2",
    "escalar":              "FFCDD2",
    "inicio":               "F5F5F5",
}
for stage, color in stage_colors.items():
    dxf = DifferentialStyle(fill=PatternFill(bgColor=color))
    rule = FormulaRule(formula=[f'$E4="{stage}"'], dxf=dxf)
    ws1.conditional_formatting.add(f"A4:O200", rule)

freeze_and_filter(ws1, 4)
ws1.row_dimensions[ROW_H].height = 40


# ══════════════════════════════════════════════════════════════
# 2. HOJA: Pacientes
# ══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Pacientes")
ws2.sheet_properties.tabColor = "16A085"
ws2.sheet_view.showGridLines = False

title_block(ws2, "👥  PACIENTES — ÍNDICE CROSS-CANAL",
            "Un registro por paciente real, identificado por teléfono — permite reconocer al mismo paciente en distintos canales", "16A085")

COLS_PAC = [
    ("patient_id",           26, "ID único  ej: PAC-1714300000-AB1CD"),
    ("telefono",             18, "Teléfono original recibido"),
    ("telefono_normalizado", 20, "9 dígitos sin prefijo — clave de búsqueda cross-canal"),
    ("nombre",               22, "Nombre completo del paciente"),
    ("canal_origen",         16, "Primer canal que usó"),
    ("canales_usados",       22, 'JSON array  ej: ["whatsapp","instagram"]'),
    ("citas_confirmadas",    18, "Contador de citas completadas"),
    ("metadata_global",      45, "JSON con datos heredables: edad, condiciones, tratamientos, aptitudVerificada"),
    ("last_activity",        22, "Última actividad (ISO timestamp)"),
]

ws2.row_dimensions[3].height = 38
for i, (col_name, width, tip) in enumerate(COLS_PAC, 1):
    bg = "16A085" if i % 2 == 1 else "0E8070"
    header_cell(ws2, 3, i, col_name, bg=bg, width=width)

sample_pac = [
    ["PAC-1714300000-AB1CD","600000001","600000001","María García","whatsapp",'["whatsapp","instagram"]',"1",'{"edad":32,"condiciones":[],"tratamientos_interes":["Botox facial"],"aptitudVerificada":true}',"2026-04-29T10:32:00Z"],
    ["PAC-1714200000-XY2ZW","610000002","610000002","Ana Martínez","facebook",'["facebook"]',"2",'{"edad":28,"condiciones":[],"tratamientos_interes":["Limpieza Facial Profunda","Mesoterapia facial"],"aptitudVerificada":true}',"2026-04-28T16:00:00Z"],
    ["PAC-1714100000-MN3PQ","655000003","655000003","Pedro Sánchez","whatsapp",'["whatsapp"]',"0",'{"edad":45,"condiciones":["anticoagulante"],"tratamientos_interes":["PRP"],"aptitudVerificada":false}',"2026-04-27T09:15:00Z"],
]

for r_idx, row in enumerate(sample_pac, 4):
    ws2.row_dimensions[r_idx].height = 36
    data_row(ws2, r_idx, row, even=(r_idx % 2 == 0))

# Highlight citas_confirmadas > 0
dxf_citas = DifferentialStyle(fill=PatternFill(bgColor="C8E6C9"), font=Font(bold=True, color="1B5E20"))
ws2.conditional_formatting.add("G4:G200", CellIsRule(operator="greaterThan", formula=["0"], dxf=dxf_citas))

freeze_and_filter(ws2, 4)


# ══════════════════════════════════════════════════════════════
# 3. HOJA: CRM_Inbox
# ══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("CRM_Inbox")
ws3.sheet_properties.tabColor = C_ACCENT
ws3.sheet_view.showGridLines = False

title_block(ws3, "📥  CRM INBOX — BANDEJA UNIFICADA",
            "Registro de todos los eventos relevantes (mensajes, citas, escalados) de todos los canales — vista operativa del staff", C_ACCENT)

COLS_CRM = [
    ("event_id",          22, "ID único del evento"),
    ("conversation_id",   28, "FK → Contexto"),
    ("patient_id",        26, "FK → Pacientes"),
    ("canal",             14, "whatsapp | instagram | facebook"),
    ("paciente_nombre",   22, "Nombre del paciente"),
    ("paciente_telefono", 18, "Teléfono"),
    ("tipo_evento",       20, "MENSAJE | ESCALADO | INTENTO_AGENDAR | CONFIRMANDO_CITA | CITA_CONFIRMADA"),
    ("prioridad",         14, "ALTA | MEDIA | BAJA"),
    ("etapa_actual",      18, "Etapa de la máquina de estados"),
    ("mensaje_paciente",  38, "Texto enviado por el paciente"),
    ("respuesta_bot",     38, "Texto respondido por Sofía"),
    ("servicio",          22, "Tratamiento mencionado"),
    ("cita_fecha",        22, "ISO timestamp de la cita agendada"),
    ("calendar_event_id", 22, "ID del evento en Google Calendar"),
    ("requiere_atencion", 16, "TRUE si necesita revisión humana"),
    ("operador_asignado", 20, "Staff que tomó el caso"),
    ("resuelto",          12, "TRUE / FALSE — el staff lo marca"),
    ("notas_staff",       35, "Notas internas del operador"),
    ("timestamp",         22, "Cuándo ocurrió el evento"),
]

ws3.row_dimensions[3].height = 42
for i, (col_name, width, tip) in enumerate(COLS_CRM, 1):
    bg = C_ACCENT if i % 2 == 1 else "C0392B"
    header_cell(ws3, 3, i, col_name, bg=bg, width=width)

sample_crm = [
    ["EVT-1714300001","whatsapp:34600000001","PAC-1714300000-AB1CD","whatsapp","María García","600000001",
     "CITA_CONFIRMADA","MEDIA","agendado","Si confirmo!","Cita confirmada María! Fecha: martes 6 de mayo a las 10:00...",
     "Botox facial","2026-05-06T10:00:00Z","cal_event_abc123","FALSE","","FALSE","","2026-04-29T10:35:00Z"],
    ["EVT-1714300002","instagram:987654321","","instagram","Carlos López","987654321",
     "ESCALADO","ALTA","escalar","Estoy tomando Sintrom","Un especialista te atenderá en breve...",
     "PRP Plasma Rico Plaquetas","","","TRUE","Pendiente","FALSE","Revisar historial médico","2026-04-29T11:06:00Z"],
    ["EVT-1714200001","facebook:111222333","PAC-1714200000-XY2ZW","facebook","Ana Martínez","610000002",
     "MENSAJE","BAJA","inicio","Hola buenas tardes","Hola Ana! Soy Sofía, tu asistente...",
     "","","","FALSE","","FALSE","","2026-04-28T15:50:00Z"],
    ["EVT-1714200002","facebook:111222333","PAC-1714200000-XY2ZW","facebook","Ana Martínez","610000002",
     "CITA_CONFIRMADA","MEDIA","agendado","si perfecto","Cita confirmada Ana!...",
     "Limpieza Facial Profunda","2026-04-30T12:00:00Z","cal_event_xyz789","FALSE","","TRUE","Cliente VIP","2026-04-28T16:00:00Z"],
]

for r_idx, row in enumerate(sample_crm, 4):
    ws3.row_dimensions[r_idx].height = 44

# Color condicional por tipo_evento (col G = 7)
evento_colors = {
    "ESCALADO":          ("FFCDD2", "B71C1C"),
    "CITA_CONFIRMADA":   ("C8E6C9", "1B5E20"),
    "INTENTO_AGENDAR":   ("E1BEE7", "4A148C"),
    "CONFIRMANDO_CITA":  ("B3E5FC", "01579B"),
    "MENSAJE":           ("F5F5F5", "424242"),
}

for r_idx, row in enumerate(sample_crm, 4):
    tipo = row[6]
    bg, fg = evento_colors.get(tipo, ("FFFFFF", "000000"))
    for col_i, val in enumerate(row, 1):
        c = ws3.cell(row=r_idx, column=col_i, value=val)
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(size=9, name="Arial", color=fg,
                      bold=(col_i in [7, 8, 15]))  # bold on tipo, prioridad, requiere_atencion
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = thin_border()

# Prioridad color (col H = 8)
dxf_alta  = DifferentialStyle(fill=PatternFill(bgColor="EF9A9A"), font=Font(bold=True, color="B71C1C"))
dxf_media = DifferentialStyle(fill=PatternFill(bgColor="FFE082"), font=Font(bold=True, color="E65100"))
dxf_baja  = DifferentialStyle(fill=PatternFill(bgColor="C8E6C9"), font=Font(color="1B5E20"))

ws3.conditional_formatting.add("H4:H500", FormulaRule(formula=['$H4="ALTA"'],  dxf=dxf_alta))
ws3.conditional_formatting.add("H4:H500", FormulaRule(formula=['$H4="MEDIA"'], dxf=dxf_media))
ws3.conditional_formatting.add("H4:H500", FormulaRule(formula=['$H4="BAJA"'],  dxf=dxf_baja))

# Resuelto = TRUE → fila gris
dxf_resuelto = DifferentialStyle(fill=PatternFill(bgColor="EEEEEE"), font=Font(color="AAAAAA"))
ws3.conditional_formatting.add("A4:S500", FormulaRule(formula=['$Q4="TRUE"'], dxf=dxf_resuelto))

# Data validation para "resuelto" y "requiere_atencion"
dv_bool = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
ws3.add_data_validation(dv_bool)
dv_bool.add("Q4:Q500")
dv_bool.add("O4:O500")

freeze_and_filter(ws3, 4)


# ══════════════════════════════════════════════════════════════
# 4. HOJA: Intervenciones
# ══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Intervenciones")
ws4.sheet_properties.tabColor = "F39C12"
ws4.sheet_view.showGridLines = False

title_block(ws4, "🎛️  INTERVENCIONES — CONTROL HUMANO",
            "El staff activa aquí la toma de control por conversación — cuando activa=TRUE el bot se silencia automáticamente", "E67E22")

# Instrucciones
ws4.row_dimensions[3].height = 14
ws4.merge_cells("A3:H3")
c_inst = ws4["A3"]
c_inst.value = "  ⚡  Para tomar control: añade una fila con el conversation_id y cambia activa a TRUE. El bot dejará de responder inmediatamente en esa conversación."
c_inst.font = Font(italic=True, size=9, color="E67E22", name="Arial")
c_inst.fill = PatternFill("solid", fgColor="FFF8E1")
c_inst.alignment = Alignment(vertical="center")

COLS_INT = [
    ("conversation_id",  30, "ID de la conversación a intervenir  ej: whatsapp:34600000001"),
    ("activa",           12, "TRUE = bot silenciado | FALSE = bot activo"),
    ("operador",         22, "Nombre del staff que interviene"),
    ("motivo",           30, "Por qué se interviene"),
    ("canal",            14, "Canal de la conversación"),
    ("paciente_nombre",  22, "Nombre del paciente (informativo)"),
    ("notas",            40, "Notas internas durante la intervención"),
    ("created_at",       22, "Cuándo se activó"),
    ("resolved_at",      22, "Cuándo se desactivó"),
]

ws4.row_dimensions[4].height = 38
for i, (col_name, width, tip) in enumerate(COLS_INT, 1):
    bg = "E67E22" if i % 2 == 1 else "D35400"
    header_cell(ws4, 4, i, col_name, bg=bg, width=width)

sample_int = [
    ["instagram:987654321","TRUE","Laura Pérez","Paciente con condición médica - revisar historial","instagram","Carlos López","Paciente toma Sintrom - derivar al Dr. Ruiz","2026-04-29T11:10:00Z",""],
    ["whatsapp:34611000099","FALSE","Miguel Torres","Paciente recurrente quería hablar con alguien","whatsapp","Rosa Díaz","Resuelta, se reprogramó la cita","2026-04-28T09:00:00Z","2026-04-28T09:30:00Z"],
]

for r_idx, row in enumerate(sample_int, 5):
    ws4.row_dimensions[r_idx].height = 36
    activa = row[1] == "TRUE"
    bg = "FFEBEE" if activa else "F1F8E9"
    fg_main = "B71C1C" if activa else "33691E"
    for col_i, val in enumerate(row, 1):
        c = ws4.cell(row=r_idx, column=col_i, value=val)
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(size=9, name="Arial", color=fg_main if col_i == 2 else "333333",
                      bold=(col_i == 2))
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = thin_border()

# Conditional: activa=TRUE → fila roja claro
dxf_activa = DifferentialStyle(fill=PatternFill(bgColor="FFCDD2"), font=Font(bold=True, color="B71C1C"))
dxf_inact  = DifferentialStyle(fill=PatternFill(bgColor="F1F8E9"), font=Font(color="33691E"))
ws4.conditional_formatting.add("A5:I200", FormulaRule(formula=['$B5="TRUE"'],  dxf=dxf_activa))
ws4.conditional_formatting.add("A5:I200", FormulaRule(formula=['$B5="FALSE"'], dxf=dxf_inact))

# Data validation para activa
dv_activa = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False, showDropDown=False)
ws4.add_data_validation(dv_activa)
dv_activa.add("B5:B200")

# Freeze at row 5
ws4.freeze_panes = ws4.cell(row=5, column=1)
ws4.auto_filter.ref = f"A4:{get_column_letter(len(COLS_INT))}4"


# ══════════════════════════════════════════════════════════════
# 5. HOJA: Dashboard (visual overview)
# ══════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Dashboard")
wb.move_sheet("Dashboard", offset=-10)  # Move to first position
ws5.sheet_properties.tabColor = "8E44AD"
ws5.sheet_view.showGridLines = False

# Big title
ws5.row_dimensions[1].height = 50
ws5.merge_cells("A1:L1")
c_title = ws5["A1"]
c_title.value = "🏥  Omnichannel AI Scheduler — Medical Clinic"
c_title.font = Font(bold=True, size=18, color="FFFFFF", name="Arial")
c_title.fill = PatternFill("solid", fgColor="1A1A2E")
c_title.alignment = Alignment(horizontal="center", vertical="center")

ws5.row_dimensions[2].height = 10
ws5.merge_cells("A2:L2")
ws5["A2"].fill = PatternFill("solid", fgColor=C_ACCENT)

# Subtitle
ws5.row_dimensions[3].height = 22
ws5.merge_cells("A3:L3")
c_sub = ws5["A3"]
c_sub.value = "  Sistema de captación y agendado 100% automatizado y omnicanal para clínicas de medicina estética"
c_sub.font = Font(italic=True, size=10, color="555555", name="Arial")
c_sub.fill = PatternFill("solid", fgColor="F8F9FA")
c_sub.alignment = Alignment(vertical="center")

ws5.row_dimensions[4].height = 14

# ── KPI Cards ──────────────────────────────────────────────────
def kpi_card(ws, start_row, start_col, title, formula, color, emoji):
    ws.row_dimensions[start_row].height = 18
    ws.row_dimensions[start_row + 1].height = 34
    ws.row_dimensions[start_row + 2].height = 16

    # Card bg
    for r in range(start_row, start_row + 3):
        for c in range(start_col, start_col + 2):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=color)

    title_c = ws.cell(row=start_row, column=start_col, value=f"  {emoji}  {title}")
    title_c.font = Font(bold=True, size=9, color="FFFFFF", name="Arial")
    title_c.fill = PatternFill("solid", fgColor=color)
    title_c.alignment = Alignment(vertical="center")
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col+1)

    val_c = ws.cell(row=start_row + 1, column=start_col, value=formula)
    val_c.font = Font(bold=True, size=22, color="FFFFFF", name="Arial")
    val_c.fill = PatternFill("solid", fgColor=color)
    val_c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=start_row+1, start_column=start_col, end_row=start_row+1, end_column=start_col+1)

kpi_data = [
    (5, 1,  "Conversaciones", '=COUNTA(Contexto!A4:A1000)', "0F3460", "💬"),
    (5, 4,  "Pacientes",      '=COUNTA(Pacientes!A4:A1000)', "16A085", "👥"),
    (5, 7,  "Citas Confirmadas", '=COUNTIF(CRM_Inbox!G4:G1000,"CITA_CONFIRMADA")', "27AE60", "✅"),
    (5, 10, "Escalados Pendientes", '=COUNTIFS(CRM_Inbox!G4:G1000,"ESCALADO",CRM_Inbox!Q4:Q1000,"FALSE")', C_ACCENT, "🚨"),
]

for sr, sc, title, formula, color, emoji in kpi_data:
    kpi_card(ws5, sr, sc, title, formula, color, emoji)
    ws5.column_dimensions[get_column_letter(sc)].width = 16
    ws5.column_dimensions[get_column_letter(sc+1)].width = 8

ws5.row_dimensions[8].height = 14

# ── Section: Hojas del sistema ─────────────────────────────────
ws5.row_dimensions[9].height = 26
ws5.merge_cells("A9:L9")
sec1 = ws5["A9"]
sec1.value = "  📊  Estructura del Spreadsheet"
sec1.font = Font(bold=True, size=11, color="FFFFFF", name="Arial")
sec1.fill = PatternFill("solid", fgColor=C_BLUE)
sec1.alignment = Alignment(vertical="center")

sheets_info = [
    ("📋 Contexto",     "Historial activo por conversación",               "Leída y escrita por el bot en cada mensaje",                  "Bot (n8n)",     "0F3460"),
    ("👥 Pacientes",    "Índice cross-canal por teléfono normalizado",      "Vincula al mismo paciente en WhatsApp, Instagram y Facebook", "Bot (n8n)",     "16A085"),
    ("📥 CRM_Inbox",   "Bandeja unificada de todos los eventos",           "ESCALADO, CITA_CONFIRMADA, MENSAJE, INTENTO_AGENDAR...",      "Bot + Staff",   C_ACCENT),
    ("🎛️ Intervenciones","Control de toma manual por el staff",             "activa=TRUE silencia el bot para esa conversación",           "Staff (manual)","E67E22"),
]

header_info_cols = ["Hoja", "Propósito", "Detalle", "Quién escribe"]
ws5.row_dimensions[10].height = 28
for ci, h in enumerate(header_info_cols, 1):
    c = ws5.cell(row=10, column=ci, value=h)
    c.font = Font(bold=True, size=9, color="FFFFFF", name="Arial")
    c.fill = PatternFill("solid", fgColor=C_MID)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border()

col_widths_info = [20, 40, 52, 18]
for ci, w in enumerate(col_widths_info, 1):
    ws5.column_dimensions[get_column_letter(ci)].width = w

for ri, (sheet, prop, detail, who, color) in enumerate(sheets_info, 11):
    ws5.row_dimensions[ri].height = 28
    for ci, val in enumerate([sheet, prop, detail, who], 1):
        c = ws5.cell(row=ri, column=ci, value=val)
        c.fill = PatternFill("solid", fgColor=color if ci == 1 else ("F8F9FA" if ri % 2 == 0 else "FFFFFF"))
        c.font = Font(size=9, name="Arial",
                      color="FFFFFF" if ci == 1 else "333333",
                      bold=(ci == 1))
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = thin_border()

ws5.row_dimensions[15].height = 14

# ── Section: Variables de entorno ─────────────────────────────
ws5.row_dimensions[16].height = 26
ws5.merge_cells("A16:L16")
sec2 = ws5["A16"]
sec2.value = "  ⚙️  Variables de Entorno Requeridas en n8n"
sec2.font = Font(bold=True, size=11, color="FFFFFF", name="Arial")
sec2.fill = PatternFill("solid", fgColor="6C3483")
sec2.alignment = Alignment(vertical="center")

env_vars = [
    ("GEMINI_API_KEY",           "Google AI Studio",           "API key para el modelo de lenguaje Gemini 2.5 Flash",             "Obligatoria"),
    ("WHATSAPP_PHONE_NUMBER_ID", "Meta for Developers",        "ID del número de teléfono de WhatsApp Business",                  "Obligatoria"),
    ("WHATSAPP_ACCESS_TOKEN",    "Meta for Developers",        "Token de acceso permanente para enviar mensajes vía WhatsApp",    "Obligatoria"),
    ("INSTAGRAM_PAGE_ID",        "Meta for Developers",        "ID de la página de Facebook vinculada a Instagram",               "Obligatoria"),
    ("FACEBOOK_PAGE_ID",         "Meta for Developers",        "ID de la página de Facebook para Messenger",                      "Obligatoria"),
    ("META_ACCESS_TOKEN",        "Meta for Developers",        "Token de página (válido para Instagram DMs y Facebook Messenger)", "Obligatoria"),
    ("META_VERIFY_TOKEN",        "Tú lo defines",              "Token secreto para verificación GET de webhooks de Meta",         "Obligatoria"),
    ("STAFF_EMAIL",              "Tu email de trabajo",        "Dirección donde llegan alertas de escalados y citas",             "Recomendada"),
    ("STAFF_WEBHOOK_URL",        "Slack / Teams / Discord",    "URL de incoming webhook para alertas instantáneas al equipo",    "Opcional"),
]

ws5.row_dimensions[17].height = 26
for ci, h in enumerate(["Variable", "Dónde obtenerla", "Descripción", "Estado"], 1):
    c = ws5.cell(row=17, column=ci, value=h)
    c.font = Font(bold=True, size=9, color="FFFFFF", name="Arial")
    c.fill = PatternFill("solid", fgColor="6C3483")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border()

for ri, (var, source, desc, status) in enumerate(env_vars, 18):
    ws5.row_dimensions[ri].height = 24
    status_color = "E8F5E9" if status == "Opcional" else ("FFF9C4" if status == "Recomendada" else "E3F2FD")
    for ci, val in enumerate([var, source, desc, status], 1):
        c = ws5.cell(row=ri, column=ci, value=val)
        bg = "F3E5F5" if ri % 2 == 0 else "FAFAFA"
        c.fill = PatternFill("solid", fgColor=status_color if ci == 4 else bg)
        c.font = Font(size=9, name="Arial", color="6C3483" if ci == 1 else "333333",
                      bold=(ci == 1), name_="Courier New" if ci == 1 else "Arial")
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = thin_border()

col_widths_env = [30, 22, 55, 14]
for ci, w in enumerate(col_widths_env, 1):
    ws5.column_dimensions[get_column_letter(ci)].width = w


# ── Section: Flujo del sistema ─────────────────────────────────
last_row = 18 + len(env_vars)
ws5.row_dimensions[last_row].height = 14
ws5.row_dimensions[last_row + 1].height = 26

ws5.merge_cells(f"A{last_row+1}:L{last_row+1}")
sec3 = ws5[f"A{last_row+1}"]
sec3.value = "  🔄  Flujo de Agendado Multi-turno (Máquina de Estados)"
sec3.font = Font(bold=True, size=11, color="FFFFFF", name="Arial")
sec3.fill = PatternFill("solid", fgColor="1A6B3C")
sec3.alignment = Alignment(vertical="center")

stages = [
    ("inicio",               "Clasificación de intención por IA",                        "27AE60"),
    ("triaje",               "¿Edad? ¿Embarazo? ¿Anticoagulantes? → apto/no apto",       "F39C12"),
    ("recopilando_nombre",   "\"¿Cuál es tu nombre completo?\"",                          "2980B9"),
    ("recopilando_telefono", "\"¿Cuál es tu teléfono de contacto?\"",                    "2980B9"),
    ("mostrando_slots",      "Muestra hasta 6 huecos libres reales del Google Calendar", "8E44AD"),
    ("confirmando_cita",     "Resumen → \"¿Confirmas? SÍ / NO\"",                        "C0392B"),
    ("agendado",             "Crea evento en Google Calendar + mensaje de confirmación",  "27AE60"),
    ("escalar",              "No apto o condición médica → alerta email + Slack al staff","E74C3C"),
]

ws5.row_dimensions[last_row + 2].height = 26
for ci, h in enumerate(["Etapa", "Qué hace el bot"], 1):
    c = ws5.cell(row=last_row+2, column=ci, value=h)
    c.font = Font(bold=True, size=9, color="FFFFFF", name="Arial")
    c.fill = PatternFill("solid", fgColor="1A6B3C")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border()

ws5.column_dimensions["A"].width = 26
ws5.column_dimensions["B"].width = 58

for ri, (stage, desc, color) in enumerate(stages, last_row + 3):
    ws5.row_dimensions[ri].height = 24
    for ci, val in enumerate([stage, desc], 1):
        c = ws5.cell(row=ri, column=ci, value=val)
        c.fill = PatternFill("solid", fgColor=color if ci == 1 else ("F8F9FA" if ri % 2 == 0 else "FFFFFF"))
        c.font = Font(size=9, name="Arial",
                      color="FFFFFF" if ci == 1 else "333333",
                      bold=(ci == 1))
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = thin_border()


# ── Save ───────────────────────────────────────────────────────
out = "/mnt/user-data/outputs/OmnichannelAI_Scheduler_GoogleSheets.xlsx"
wb.save(out)
print(f"OK: {out}")
print(f"Hojas: {wb.sheetnames}")
